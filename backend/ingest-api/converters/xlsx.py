from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def extract_xlsx(file_bytes: bytes) -> tuple[str, dict]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sections: list[str] = []
    for sheet in wb.worksheets:
        sections.append(f"## {sheet.title}")
        merged_map = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = sheet.cell(min_row, min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[(row, col)] = value
        rows = []
        for row in sheet.iter_rows():
            values = []
            for cell in row:
                value = merged_map.get((cell.row, cell.column), cell.value)
                values.append("" if value is None else str(value))
            if any(values):
                rows.append(values)
        if not rows:
            continue
        header = "| " + " | ".join(rows[0]) + " |"
        divider = "| " + " | ".join(["---"] * len(rows[0])) + " |"
        body = ["| " + " | ".join(r) + " |" for r in rows[1:]]
        sections.append("\n".join([header, divider, *body]))
    return "\n\n".join(sections).strip(), {"sheet_count": len(wb.worksheets)}
